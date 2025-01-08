import pdfplumber
import pandas as pd
from difflib import SequenceMatcher
import os
from datetime import datetime

from openpyxl import load_workbook # read excel file with openpyxl to preserve formatting
import glob

# split() separates the words in a line
# strip() removes extra whitespaces

## TRY TO HANDLE MULTIPLE PDFs AT A TIME!!!
## TRY TO HANDLE PAGE 2 IF RM IS NOT FOUND IN PAGE 1

def clean_text(text):
    # print("Normalizing text for comparison...")

    text = text.lower()

    # print("Remove common words that don't add meaning...")
    stop_words = ['to', 'at', 'and', 'with', 'for', 'the', '-', 'service', 'power', 'plant']
    for word in stop_words:
        text = text.replace(f' {word} ', ' ')


    # Remove numbers and special characters
    text = ''.join(char for char in text if not char.isdigit())
    text = ''.join(char if char.isalnum() or char.isspace() else ' ' for char in text)


    # Remove extra spaces
    text = ' '.join(text.split())
    print(f"Cleaned remark:", text)
    
    return text

## extract text from pdf using pdfplumber
## since all data available from the first page, we get page 1 only
def extract_text_from_pdf(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            first_page = pdf.pages[0]

            text = first_page.extract_text()

            # if the 'Amount' can't be found on Page 1
            amount_found = 'Total (MYR)' in text

            # only extract page 2 if amount cant be found
            if not amount_found and len(pdf.pages) > 1:
                print("Amount not found in first page, checking second page...")
                second_page = pdf.pages[1]
                second_page_text = second_page.extract_text()
                text = text + "\n<PAGE_BREAK>\n" + second_page_text

            return text

    except Exception as e:
        print(f"Text can't be extracted from page 1.")

        return None
    
def extract_po_from_pdf(pdf_text):
    print("Extracting Purchase Order (PO) information from PDF...")

    """
    'GL' = ['Routine Maintenance', 'Çalibration', 'Sampling Test', /
    'Tools & Consumables', 'Stationery & Printing', 'Freight, /
    Cartage and Courier', 'Staff Welfare and Entertainment]

    """

    po_data = {
        'PO No': None,
        'Date': None,
        'Vendor': None,
        'Amount': None,
        'GL': 'Insert Category',
        'Month': None
    }

    if not pdf_text:
        return po_data
    
    # split the extracted pdf into individual text
    lines = pdf_text.split('\n')
    description = [] # collect the lines of description
    line_count = 0
    found_marker = False # track if 'MYR' line is found

    for i, line in enumerate(lines):
        # print(f"Processing line {i+1}: {line}")  # Debug print
        
        if 'PURCHASE ORDER NO' in line:
            # extract the last word in the line
            po_data['PO No'] = line.split()[-1].strip()

        if 'ORDER DATE' in line:
            # Split the line and get everything after 'ORDER DATE'
            date_str = ' '.join(line.split('ORDER DATE')[1].strip().split())

            # convert date format from e.g. October 3, 2024 to 10/3/2024
            
            # first, create datetime object
            date_obj = datetime.strptime(date_str, '%B %d, %Y')

            po_data['Date'] = date_obj.strftime('%m/%d/%Y')
            po_data['Month'] = date_obj.strftime('%B')

            po_data['Vendor'] = line.split('ORDER DATE')[0].strip()

        if 'Total (MYR)' in line:
            po_data['Amount'] = 'RM' + line.split()[-1].strip()

        # finding the description lines 
        if 'MYR MYR MYR MYR' in line:
            found_marker = True
            continue # Skip the marker line itself

        if found_marker and line.strip() and line_count < 3: # takes up to 3 lines after
            description.append(line.strip())
            line_count += 1


    # combine the description lines into one line
    description = ' '.join(description)
    description = description.strip()

    return po_data, description


def update_excel_with_po(excel_path, po_data, po_description):
    print("Reading excel file...")

    wb = load_workbook(excel_path)
    ws = wb.active

    # first, read with pandas to find the matching description
    df = pd.read_excel(excel_path)

    best_match_idx = -1 # find matching row index
    highest_similarity = 0

    # clean both po description and pr remark
    po_description_clean = clean_text(po_description)
    print(f"Cleaned description:", po_description_clean)

    for idx, row in df.iterrows():
        print("Matching PO description and PR Remark...")

        # skip if Remark is NaN/empty, since it is type 'float'
        if pd.isna(row['Remark']):
            continue

        remark_clean = clean_text(row['Remark']) # clean the remark column


        similarity = SequenceMatcher(None, po_description_clean, remark_clean).ratio()
        print(f"Similarity with row {idx + 1}: {similarity}")  # Debug print
        
        if similarity > highest_similarity:
            highest_similarity = similarity
            best_match_idx = idx

    # Update matching row with PO data
    if best_match_idx >= 0:

        excel_row = best_match_idx + 2 # add 1 for header index, add 1 for 1-based indexing
        
        # openpyxl reads column as A, B, C, D, not as column names like pandas
        po_columns = {
            'PO No': 'F',
            'Date': 'G', 
            'Vendor': 'H',
            'Amount': 'I',
            'GL': 'J',
            'Month': 'K'
        }

        # check is any PO cell is filled
        has_existing_po = any(ws[f'{col}{excel_row}'].value is not None 
                            for col in po_columns.values())

        if has_existing_po:
            print(f"Row {best_match_idx + 1} already contains PO data. Skipping update.")
            return None

        # If we get here, all PO cells are empty, safe to update
        for field, col in po_columns.items():
            cell = ws[f'{col}{excel_row}']
            # Preserve cell format but update value
            cell.value = po_data[field]

        print("Match found! Updating existing file...")
        
        # Save updated file
        wb.save(excel_path)
        return df
    
    print("No match found") #else
    
    return None

def main(): # main function

    ### NAME OF THE FOLDER WHERE PDF FILES ARE STORED
    pdf_folder = "PO_in_excel/"

    ### PATH OF EXCEL TO BE UPDATED
    excel_path = "Trial (in excel).xlsx"

    # find files with pdf format
    pdf_files = glob.glob(pdf_folder + "*.pdf")
    print(f"Found {len(pdf_files)} PDF files")

    # Process each PDF
    for pdf_file in pdf_files:
        print(f"\nProcessing {pdf_file}...")
        
        try:
            # Extract text from PDF
            text = extract_text_from_pdf(pdf_file)
            # print(text)

            if text is None:
                print(f"Could not extract text from {pdf_file}")
                continue
                
            # Get PO data
            po_data, description = extract_po_from_pdf(text)
            print("Extracted PO Data:", po_data)
            print("PO Description:", description)
            
            # Update Excel
            result = update_excel_with_po(excel_path, po_data, description)
            
            if result is not None:
                print(f"Successfully updated Excel with data from {pdf_file}")
            else:
                print(f"No matching row found for {pdf_file}")
                
        except Exception as e:
            print(f"Error processing {pdf_file}: {str(e)}")
            continue
    
    print("\nFinished processing all PDFs")


if __name__ == "__main__":
    main()



